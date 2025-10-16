import os
import re
import warnings
from pathlib import Path
from typing import Dict, Optional, Tuple, Union, List
from concurrent.futures import ThreadPoolExecutor, as_completed  # dùng thread để tránh lỗi spawn trên Windows
from datetime import datetime

# Ẩn cảnh báo Data Validation của openpyxl cho sạch log
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl",
    message=".*Data Validation extension is not supported.*",
)

from openpyxl import load_workbook

# =========================
# ====== CẤU HÌNH =========
# =========================
FE_FOLDER   = r"D:\Documents\027-社内レビュー\モック開発\海外\ベトナムチーム"
BE_FOLDER   = r"D:\Documents\027-社内レビュー\WEBAPI開発\海外\ベトナム"
TC_FOLDER   = r"D:\Documents\90.作業情報\50.作業報告\KMD（ケアマネ先行開発）\成果物\単体テスト仕様"
EXEC_FOLDER = r"D:\Documents\90.作業情報\50.作業報告\KMD（ケアマネ先行開発）\成果物\単体テスト実施"

# ★ File chứa danh sách màn hình mục tiêu
SUMMARY_FILE = r"C:\Users\KDVN-ANHNC\Documents\tool\SoukeiTool\画面_summary.xlsx"

OUTPUT_FILE    = "Screen_LOC_Summary.xlsx"
ERROR_LOG_FILE = "collect_and_fill_error_log.txt"  # TXT log lỗi (cùng thư mục script)

# Hiệu năng
MAX_WORKERS    = max(2, (os.cpu_count() or 4))
PROGRESS_EVERY = 25
VERBOSE        = False
USE_THREADS    = True  # đặt True để dùng ThreadPool (ổn định trên Windows)

def vprint(*args, **kwargs):
    if VERBOSE:
        print(*args, **kwargs)

# =========================
# ====== HÀM PHỤ ==========

ERROR_STRINGS = {"#REF!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#N/A", "#VALUE!"}
GUI_RE = re.compile(r"(GUI\d{5})")

def normalize_name(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    s = str(s).strip().replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s)
    return s

def is_error_value(val) -> bool:
    return isinstance(val, str) and val.strip().upper() in ERROR_STRINGS

def coerce_number(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip()
        if is_error_value(t):
            return None
        m = re.search(r"[-+]?\d[\d,\.]*", t)
        if not m:
            return None
        t = m.group(0).replace(",", "")
        try:
            return float(t)
        except Exception:
            return None
    try:
        return float(v)
    except Exception:
        return None

def list_excel_files(folder: str):
    """Duyệt đệ quy .xlsx/.xlsm, bỏ qua file tạm ~$*."""
    p = Path(folder)
    for f in p.rglob("*"):
        if (
            f.is_file()
            and f.suffix.lower() in {".xlsx", ".xlsm"}
            and not f.name.startswith("~$")
        ):
            yield str(f)

def try_pick_sheet_name(path: str, preferred: str | None):
    """Ưu tiên sheet preferred (khớp chính xác hoặc chứa), nếu không có thì lấy active."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        sheets = wb.sheetnames
        if preferred:
            if preferred in sheets:
                name = preferred
            else:
                lowmap = {sn.lower(): sn for sn in sheets}
                match = next((orig for low, orig in lowmap.items() if preferred.lower() in low), None)
                name = match if match else sheets[0]
        else:
            name = sheets[0]
        wb.close()
        return name
    except Exception:
        return preferred or "Sheet1"

def get_cell_value_fast(path: str, sheet_name: str, addr: str):
    """
    Đọc cell nhanh (read_only=True, data_only=True).
    Nếu None → mở lại data_only=False (read_only=True) để xem lỗi/công thức.
    Ưu tiên trả '#REF!' nếu phát hiện.
    """
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        val = ws[addr].value
        wb.close()
        if is_error_value(val) or val is not None:
            return val

        wb2 = load_workbook(path, data_only=False, read_only=True)
        ws2 = wb2[sheet_name] if sheet_name in wb2.sheetnames else wb2.active
        val2 = ws2[addr].value
        wb2.close()
        if isinstance(val2, str) and ("#REF!" in val2 or is_error_value(val2)):
            return "#REF!"
        return val2
    except Exception:
        return None

def extract_gui_from_filename(path: str) -> Optional[str]:
    m = GUI_RE.search(Path(path).name)
    return m.group(1) if m else None

def read_name_and_loc(path: str) -> Tuple[Optional[str], Union[str, float, None]]:
    """Đọc G5 (Tên màn hình) và LOC từ AU6 (→ AV6 nếu rỗng), ưu tiên sheet レビュー依頼書兼報告書."""
    sheet   = try_pick_sheet_name(path, "レビュー依頼書兼報告書")
    name_v  = get_cell_value_fast(path, sheet, "G5")
    name    = normalize_name(name_v)
    au6     = get_cell_value_fast(path, sheet, "AU6")
    av6     = None if au6 is not None else get_cell_value_fast(path, sheet, "AV6")

    if is_error_value(au6):
        loc = "#REF!"
    elif au6 is not None:
        loc = coerce_number(au6) if coerce_number(au6) is not None else str(au6)
    elif is_error_value(av6):
        loc = "#REF!"
    elif av6 is not None:
        loc = coerce_number(av6) if coerce_number(av6) is not None else str(av6)
    else:
        loc = None

    vprint(f"[{Path(path).name}] sheet={sheet} G5={name_v} AU6={au6} AV6={av6} -> LOC={loc}")
    return name, loc

def extract_unique_number_from_row_values(row_vals: List[object]) -> Optional[float]:
    nums = []
    for val in row_vals:
        n = coerce_number(val)
        if n is not None:
            nums.append(float(n))
    uniq = set(nums)
    if len(uniq) == 1:
        return uniq.pop()
    return None

def find_total_cases_fast(path: str, sheet_name: str, max_scan_rows: int = 300) -> Optional[float]:
    """Tìm hàng có '総ケース数' & trả về số duy nhất trên hàng (nhanh & dừng sớm)."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        scanned = 0
        for row in ws.iter_rows(values_only=True):
            scanned += 1
            if scanned > max_scan_rows:
                break
            if not row:
                continue
            if any(isinstance(v, str) and "総ケース数" in v for v in row):
                n = extract_unique_number_from_row_values(list(row))
                wb.close()
                return n
        wb.close()
        return None
    except Exception:
        return None

def read_TC_from_spec(path: str) -> Tuple[Optional[str], Union[str, float, None]]:
    """Đọc 改訂履歴!AG1 làm tên màn; TestCase ưu tiên '総ケース数' ở 評価項目サマリ, fallback F5."""
    rev_sheet = try_pick_sheet_name(path, "改訂履歴")
    sum_sheet = try_pick_sheet_name(path, "評価項目サマリ")

    screen_raw  = get_cell_value_fast(path, rev_sheet, "AG1")
    screen_name = normalize_name(screen_raw) if screen_raw else extract_gui_from_filename(path)

    test_case = None
    if sum_sheet:
        tc = find_total_cases_fast(path, sum_sheet, max_scan_rows=300)
        if tc is not None:
            test_case = tc
        else:
            f5 = get_cell_value_fast(path, sum_sheet, "F5")
            if is_error_value(f5):
                test_case = "#REF!"
            elif f5 is not None:
                n = coerce_number(f5)
                test_case = n if n is not None else str(f5)
            else:
                test_case = None

    vprint(f"[{Path(path).name}] AG1={screen_name} 総ケース数/F5 -> {test_case}")
    return screen_name, test_case

# -------- Bước 4: NGケース数 (cộng 5 ô bên dưới, chỉ tìm tiêu đề trong hàng 1..7) --------
def find_ng_sum_fast(path: str, sheet_name: str, header_rows: int = 7) -> Optional[float]:
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        end_header_row = min(header_rows, max_row)

        hit_r = hit_c = None
        for r in range(1, end_header_row + 1):
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, str) and "NGケース数" in val:
                    hit_r, hit_c = r, c
                    break
            if hit_r is not None:
                break

        if hit_r is None:
            wb.close()
            return None

        total = 0.0
        for i in range(1, 6):
            rr = hit_r + i
            if rr > max_row:
                break
            v = ws.cell(row=rr, column=hit_c).value
            n = coerce_number(v)
            if n is not None:
                total += float(n)
        wb.close()
        return total
    except Exception:
        return None

def read_exec_ngsum(path: str) -> Tuple[Optional[str], Optional[float]]:
    """Tên màn (AG1 | GUIxxxxx từ tên file) + tổng 'NGケース数' (5 ô dưới) ở 評価項目サマリ."""
    rev_sheet = try_pick_sheet_name(path, "改訂履歴")
    sum_sheet = try_pick_sheet_name(path, "評価項目サマリ")

    screen_raw  = get_cell_value_fast(path, rev_sheet, "AG1")
    screen_name = normalize_name(screen_raw) if screen_raw else extract_gui_from_filename(path)
    if not sum_sheet:
        return screen_name, None

    ng_sum = find_ng_sum_fast(path, sum_sheet, header_rows=7)
    vprint(f"[{Path(path).name}] AG1={screen_name} NGケース数_5行合計 -> {ng_sum}")
    return screen_name, ng_sum

def safe_number_or_keep(x):
    if isinstance(x, str) and is_error_value(x):
        return x
    n = coerce_number(x)
    return n if n is not None else x

# ====== Đọc danh sách màn hình mục tiêu từ SUMMARY_FILE ======
def load_target_screens(summary_path: str) -> List[str]:
    """
    Đọc cột 'Tên màn hình' từ file summary (xlsx).
    Trả về danh sách GUIxxxxx (đã normalize, bỏ giá trị rỗng).
    """
    try:
        import pandas as pd  # lazy import để tránh import trong worker
        df = pd.read_excel(summary_path)
        col = None
        for cand in ["Tên màn hình", "Ten man hinh", "Screen", "ScreenName", "画面", "画面ID"]:
            if cand in df.columns:
                col = cand
                break
        if col is None:
            col = df.columns[0]
        screens = [normalize_name(x) for x in df[col].tolist()]
        screens = [s for s in screens if s]
        return screens
    except Exception as e:
        print(f"[WARN] Không đọc được SUMMARY_FILE: {summary_path} -> {e}")
        return []

# =========================
# ====== WORKERS ==========
def _worker_fe(path: str):
    try:
        name, loc = read_name_and_loc(path)
        if name is None and loc is None:
            return path, (name, loc), "FE: Không đọc được G5/AU6/AV6"
        return path, (name, loc), None
    except Exception as e:
        return path, (None, None), f"FE EXC: {type(e).__name__}: {e}"

def _worker_be(path: str):
    try:
        name, loc = read_name_and_loc(path)
        if name is None and loc is None:
            return path, (name, loc), "BE: Không đọc được G5/AU6/AV6"
        return path, (name, loc), None
    except Exception as e:
        return path, (None, None), f"BE EXC: {type(e).__name__}: {e}"

def _worker_tc(path: str):
    try:
        screen, tc = read_TC_from_spec(path)
        if screen is None and tc is None:
            return path, (screen, tc), "TC: Không đọc được AG1/総ケース数/F5"
        return path, (screen, tc), None
    except Exception as e:
        return path, (None, None), f"TC EXC: {type(e).__name__}: {e}"

def _worker_exec(path: str):
    try:
        screen, ngsum = read_exec_ngsum(path)
        if screen is None and ngsum is None:
            return path, (screen, ngsum), "EXEC: Không đọc được AG1/NGケース数"
        return path, (screen, ngsum), None
    except Exception as e:
        return path, (None, None), f"EXEC EXC: {type(e).__name__}: {e}"

# =========================
# ====== LUỒNG CHÍNH ======
def main():
    result: Dict[str, Dict[str, Union[float, str, None]]] = {}
    errors: List[str] = []

    # 1) Đọc danh sách màn hình mục tiêu từ SUMMARY_FILE
    target_screens = load_target_screens(SUMMARY_FILE)
    target_set = set(target_screens)
    if not target_screens:
        print("❌ Không có danh sách 'Tên màn hình' trong SUMMARY_FILE — dừng.")
        return
    print(f"🎯 Số màn hình mục tiêu: {len(target_screens)}")

    # Khởi tạo result cho tất cả màn mục tiêu
    for scr in target_screens:
        result[scr] = {"LOCFE": None, "LOCBE": None, "TestCase": None, "NGケース数_5行合計": None}

    # Xóa log cũ (nếu có)
    try:
        Path(ERROR_LOG_FILE).unlink(missing_ok=True)
    except Exception:
        pass

    Exec = ThreadPoolExecutor  # có thể thay bằng ProcessPoolExecutor nếu muốn

    # --- FE ---
    fe_files = list(list_excel_files(FE_FOLDER))
    print("=== Bước 1: FE (song song) ===")
    print(f"  -> FE files: {len(fe_files)}")
    dup_fe = 0
    processed = 0
    with Exec(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(_worker_fe, f): f for f in fe_files}
        for fut in as_completed(futures):
            processed += 1
            if processed % PROGRESS_EVERY == 0:
                print(f"    ... FE tiến độ: {processed}/{len(fe_files)}")
            path, (name, locfe), err = fut.result()
            if err:
                errors.append(f"[FE] {path} -> {err}")
            if not name or name not in target_set:
                continue
            if result[name].get("LOCFE") is not None:
                dup_fe += 1
                continue
            result[name]["LOCFE"] = locfe
    print(f"  -> FE trùng bỏ qua: {dup_fe}")

    # --- BE ---
    be_files = list(list_excel_files(BE_FOLDER))
    print("=== Bước 2: BE (song song) ===")
    print(f"  -> BE files: {len(be_files)}")
    match_be = dup_be = 0
    processed = 0
    with Exec(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(_worker_be, f): f for f in be_files}
        for fut in as_completed(futures):
            processed += 1
            if processed % PROGRESS_EVERY == 0:
                print(f"    ... BE tiến độ: {processed}/{len(be_files)}")
            path, (name, locbe), err = fut.result()
            if err:
                errors.append(f"[BE] {path} -> {err}")
            if not name or name not in target_set:
                continue
            if result[name].get("LOCBE") is None:
                result[name]["LOCBE"] = locbe
                match_be += 1
            else:
                dup_be += 1
    print(f"  -> BE matched: {match_be} | BE trùng: {dup_be}")

    # --- TC ---
    tc_files = list(list_excel_files(TC_FOLDER))
    print("=== Bước 3: 単体テスト仕様 (song song) ===")
    print(f"  -> Spec files: {len(tc_files)}")
    match_tc = 0
    processed = 0
    with Exec(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(_worker_tc, f): f for f in tc_files}
        for fut in as_completed(futures):
            processed += 1
            if processed % PROGRESS_EVERY == 0:
                print(f"    ... TC tiến độ: {processed}/{len(tc_files)}")
            path, (scr, tc), err = fut.result()
            if err:
                errors.append(f"[TC] {path} -> {err}")
            if not scr or scr not in target_set:
                continue
            result[scr]["TestCase"] = tc
            match_tc += 1
    print(f"  -> TestCase matched: {match_tc}")

    # --- EXEC (Bước 4) ---
    exec_files = list(list_excel_files(EXEC_FOLDER))
    print("=== Bước 4: 単体テスト実施 (song song) — NGケース数_5行合計 ===")
    print(f"  -> Exec files: {len(exec_files)}")
    match_exec = 0
    processed = 0
    with Exec(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(_worker_exec, f): f for f in exec_files}
        for fut in as_completed(futures):
            processed += 1
            if processed % PROGRESS_EVERY == 0:
                print(f"    ... EXEC tiến độ: {processed}/{len(exec_files)}")
            path, (scr, ngsum), err = fut.result()
            if err:
                errors.append(f"[EXEC] {path} -> {err}")
            if not scr or scr not in target_set:
                continue
            result[scr]["NGケース数_5行合計"] = ngsum
            match_exec += 1
    print(f"  -> EXEC matched: {match_exec}")

    # Xuất Excel
    print("=== Xuất Excel ===")
    import pandas as pd  # lazy import: chỉ ở tiến trình chính
    rows = []
    for name in sorted(result.keys()):
        vals = result[name]
        rows.append(
            {
                "Tên màn hình": name,
                "LOCFE": vals.get("LOCFE"),
                "LOCBE": vals.get("LOCBE"),
                "TestCase": vals.get("TestCase"),
                "NGケース数_5行合計": vals.get("NGケース数_5行合計"),
            }
        )
    df = pd.DataFrame(rows)
    for c in ["LOCFE", "LOCBE", "TestCase", "NGケース数_5行合計"]:
        df[c] = df[c].apply(safe_number_or_keep)
    df.to_excel(OUTPUT_FILE, index=False)
    print(f"✅ Đã tạo: {OUTPUT_FILE}")

    # Ghi log lỗi ra TXT (1 lần, ở tiến trình chính)
    if errors:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(ERROR_LOG_FILE, "w", encoding="utf-8") as f:
            f.write(f"[collect_and_fill] Error log — {ts}\n")
            f.write("=" * 80 + "\n")
            for line in errors:
                f.write(line + "\n")
        print(f"⚠️  Có {len(errors)} mục lỗi. Xem file log: {ERROR_LOG_FILE}")
    else:
        try:
            Path(ERROR_LOG_FILE).unlink(missing_ok=True)
        except Exception:
            pass

    print(
        f"📊 Mục tiêu: {len(result)} | BE matched: {match_be} | TC matched: {match_tc} | EXEC matched: {match_exec} | FE dup bỏ qua: {dup_fe}"
    )

if __name__ == "__main__":
    main()
